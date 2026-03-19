from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    SalesInvoice,
    SalesReturn,
    PurchaseInvoice,
    PurchaseReturn,
    Product,
    Payment,
    Party,
    Account,
)


def _fmt_date(val):
    if not val:
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _auto_width(ws, max_cols=40):
    for col in range(1, min(ws.max_column, max_cols) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18


def _wb_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# SALES LEDGER
# =========================
def generate_sales_ledger(owner) -> bytes:
    """
    Uses your actual fields:
    SalesInvoice + SalesReturn
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Ledger"

    headers = [
        "Entry Type",
        "Reference No",
        "Date",
        "Customer",
        "Customer Phone",
        "Reference Invoice",
        "Total Amount",
        "Posted",
        "Payment Type",
        "Payment Account",
        "Payment Amount",
        "Notes",
    ]
    ws.append(headers)

    rows = []

    invoice_qs = SalesInvoice.objects.filter(owner=owner).select_related("customer", "payment_account").order_by("invoice_date", "id")

    for inv in invoice_qs:
        cust = inv.customer
        cust_name = cust.name if cust else ""
        cust_phone = getattr(cust, "phone", "") if cust else ""

        acct = inv.payment_account
        acct_name = acct.name if acct else ""

        rows.append((
            inv.invoice_date,
            inv.id,
            [
            "Invoice",
            inv.invoice_number or "",
            _fmt_date(inv.invoice_date),
            cust_name,
            cust_phone,
            "",
            inv.calculate_total(),
            "YES" if inv.posted else "NO",
            inv.payment_type or "",
            acct_name,
            inv.payment_amount or 0,
            inv.notes or "",
            ],
        ))

    return_qs = SalesReturn.objects.filter(owner=owner).select_related("customer", "reference_invoice").order_by("return_date", "id")

    for ret in return_qs:
        cust = ret.customer
        cust_name = cust.name if cust else ""
        cust_phone = getattr(cust, "phone", "") if cust else ""
        ref_invoice = ret.reference_invoice

        rows.append((
            ret.return_date,
            ret.id,
            [
            "Return",
            f"SR-{ret.id}",
            _fmt_date(ret.return_date),
            cust_name,
            cust_phone,
            ref_invoice.invoice_number if ref_invoice else "",
            ret.calculate_total(),
            "YES" if ret.posted else "NO",
            "",
            "",
            0,
            ret.notes or "",
            ],
        ))

    for _, _, row in sorted(rows, key=lambda item: (item[0] or date.min, item[1])):
        ws.append(row)

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# PURCHASE LEDGER
# =========================
def generate_purchase_ledger(owner) -> bytes:
    """
    Uses your actual fields:
    PurchaseInvoice + PurchaseReturn
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Ledger"

    headers = [
        "Entry Type",
        "Reference No",
        "Date",
        "Supplier",
        "Supplier Phone",
        "Reference Invoice",
        "Total Amount",
        "Date Received",
        "Freight Charges",
        "Other Charges",
        "Posted",
        "Payment Type",
        "Payment Account",
        "Payment Amount",
        "Notes",
    ]
    ws.append(headers)

    rows = []

    invoice_qs = PurchaseInvoice.objects.filter(owner=owner).select_related("supplier", "payment_account").order_by("invoice_date", "id")

    for inv in invoice_qs:
        sup = inv.supplier
        sup_name = sup.name if sup else ""
        sup_phone = getattr(sup, "phone", "") if sup else ""

        acct = inv.payment_account
        acct_name = acct.name if acct else ""

        rows.append((
            inv.invoice_date,
            inv.id,
            [
            "Invoice",
            inv.invoice_number or "",
            _fmt_date(inv.invoice_date),
            sup_name,
            sup_phone,
            "",
            inv.calculate_total(),
            _fmt_date(inv.date_received),
            inv.freight_charges or 0,
            inv.other_charges or 0,
            "YES" if inv.posted else "NO",
            inv.payment_type or "",
            acct_name,
            inv.payment_amount or 0,
            inv.notes or "",
            ],
        ))

    return_qs = PurchaseReturn.objects.filter(owner=owner).select_related("supplier", "reference_invoice").order_by("return_date", "id")

    for ret in return_qs:
        sup = ret.supplier
        sup_name = sup.name if sup else ""
        sup_phone = getattr(sup, "phone", "") if sup else ""
        ref_invoice = ret.reference_invoice

        rows.append((
            ret.return_date,
            ret.id,
            [
            "Return",
            f"PR-{ret.id}",
            _fmt_date(ret.return_date),
            sup_name,
            sup_phone,
            ref_invoice.invoice_number if ref_invoice else "",
            ret.calculate_total(),
            "",
            0,
            0,
            "YES" if ret.posted else "NO",
            "",
            "",
            0,
            ret.notes or "",
            ],
        ))

    for _, _, row in sorted(rows, key=lambda item: (item[0] or date.min, item[1])):
        ws.append(row)

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# PAYMENTS LEDGER
# =========================
def generate_payments_ledger(owner) -> bytes:
    """
    Uses your actual fields:
    Payment: date, party, account, direction, amount, description, posted, related_model, related_id, is_adjustment, adjustment_side
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Ledger"

    headers = [
        "Date",
        "Direction",
        "Party",
        "Party Phone",
        "Account",
        "Amount",
        "Posted",
        "Description",
        "Related Model",
        "Related ID",
        "Is Adjustment",
        "Adjustment Side",
    ]
    ws.append(headers)

    qs = Payment.objects.filter(owner=owner).select_related("party", "account").order_by("date", "id")

    for p in qs:
        party = p.party
        party_name = party.name if party else ""
        party_phone = getattr(party, "phone", "") if party else ""

        acct = p.account
        acct_name = acct.name if acct else ""

        ws.append([
            _fmt_date(p.date),
            p.direction or "",
            party_name,
            party_phone,
            acct_name,
            p.amount or 0,
            "YES" if p.posted else "NO",
            p.description or "",
            p.related_model or "",
            p.related_id or "",
            "YES" if p.is_adjustment else "NO",
            p.adjustment_side or "",
        ])

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# PRODUCTS LIST
# =========================
def generate_products_list(owner) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["Code", "Product", "Unit", "Purchase Price", "Sale Price", "Current Stock", "Active"]
    ws.append(headers)

    qs = Product.objects.filter(owner=owner).order_by("id")
    for pr in qs:
        ws.append([
            pr.code or "",
            pr.name or "",
            pr.unit or "",
            pr.purchase_price_per_unit or 0,
            pr.sale_price_per_unit or 0,
            pr.current_stock or 0,
            "YES" if pr.is_active else "NO",
        ])

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# PARTIES LIST
# =========================
def generate_parties_list(owner) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Parties"

    headers = ["Party", "Type", "Phone", "City", "Address"]
    ws.append(headers)

    qs = Party.objects.filter(owner=owner).order_by("id")
    for pt in qs:
        ws.append([
            getattr(pt, "name", "") or "",
            getattr(pt, "party_type", "") or getattr(pt, "type", "") or "",
            getattr(pt, "phone", "") or "",
            getattr(pt, "city", "") or "",
            getattr(pt, "address", "") or "",
        ])

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# ACCOUNTS LIST
# =========================
def generate_accounts_list(owner) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    headers = ["Code", "Account", "Type", "Cash/Bank", "Allow for Payments"]
    ws.append(headers)

    qs = Account.objects.filter(owner=owner).order_by("id")
    for a in qs:
        ws.append([
            a.code or "",
            a.name or "",
            a.account_type or "",
            "YES" if a.is_cash_or_bank else "NO",
            "YES" if a.allow_for_payments else "NO",
        ])

    _auto_width(ws)
    return _wb_to_bytes(wb)


# =========================
# FULL ZIP
# =========================
def build_tax_pack_zip(owner) -> bytes:
    files = {
        "Sales_Ledger.xlsx": generate_sales_ledger(owner),
        "Purchase_Ledger.xlsx": generate_purchase_ledger(owner),
        "Payments_Ledger.xlsx": generate_payments_ledger(owner),
        "Products_List.xlsx": generate_products_list(owner),
        "Parties_List.xlsx": generate_parties_list(owner),
        "Accounts_List.xlsx": generate_accounts_list(owner),
    }

    bio = BytesIO()
    with ZipFile(bio, "w", compression=ZIP_DEFLATED) as zf:
        for filename, content in files.items():
            zf.writestr(filename, content)

    return bio.getvalue()
